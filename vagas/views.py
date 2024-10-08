# PARA AS VIEWS
import calendar
import json
from django.views.decorators.clickjacking import xframe_options_exempt
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
# AUTH
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.utils.timezone import make_aware
# MODELS E FORMS
from .forms import *
from django.contrib.auth.models import User
from django.db.models import Q
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
# OUTROS
from django.http import FileResponse, Http404, JsonResponse
import requests
import pdfkit
from datetime import date, datetime
# VIEWS
from django.db.models import Sum, Count
from django.utils import timezone
from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Alignment
from urllib.parse import quote

from .models import Slide
from django.http import HttpResponseForbidden, HttpResponse

from autenticacao.models import Pessoa
from django.shortcuts import get_object_or_404


import os
import subprocess
from django.http import HttpResponse
from django.conf import settings
from balcao_de_emprego.settings import db_name, db_user, db_host, db_passwd
from django.views import View

class BackupDatabaseView(View):
    def get(self, request):
        # Caminho para salvar o backup localmente
        backup_file_path = os.path.join(settings.MEDIA_ROOT, f'{db_name}_backup.sql')

        command = [
            'mysqldump',
            '-h', db_host,
            '-P', '3306',
            '-u', db_user,
            f'--password={db_passwd}',
            db_name
        ]

        try:
            # Executando o comando e salvando o backup no arquivo
            with open(backup_file_path, 'w') as backup_file:
                subprocess.run(command, stdout=backup_file, check=True)

            # Retornar o arquivo de backup como resposta para download
            with open(backup_file_path, 'rb') as backup_file:
                response = HttpResponse(backup_file.read(), content_type='application/sql')
                response['Content-Disposition'] = f'attachment; filename={os.path.basename(backup_file_path)}'
                return response

        except subprocess.CalledProcessError as e:
            # Lidar com erros durante o backup
            return HttpResponse(f"Erro ao criar backup: {str(e)}", status=500)
        
def staff_required(view_func):
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_staff:
            return HttpResponseForbidden("Acesso negado. Você não é um funcionário.")
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def superuser_required(view_func):
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_superuser:
            return HttpResponseForbidden("Acesso negado. Você não é um super usuário.")
        return view_func(request, *args, **kwargs)
    return _wrapped_view

@login_required
@staff_required
@superuser_required
def exportar_vagas_excel(request):
    # Filtrar os registros onde ativo=True
    vagas_ativas = Vaga_Emprego.objects.filter(ativo=True)

    # Criar um objeto Workbook
    wb = Workbook()
    ws = wb.active

    # Adicionar cabeçalhos das colunas
    ws.append(['ID da Vaga', 'Nome da Empresa', 'Nome do Cargo', 'Número de Vagas', 'Data de Inclusão', 'Telefone', 'Whatsapp', 'Email'])

    # Preencher os dados
    for vaga in vagas_ativas:
        ws.append([vaga.id, vaga.empresa.nome, vaga.cargo.nome, vaga.quantidadeVagas, vaga.dt_inclusao, vaga.empresa.telefone, vaga.empresa.whatsapp, vaga.empresa.email])

    # Criar uma resposta HTTP
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename={quote("vagas_ativas.xlsx")}'

    # Salvar o conteúdo do arquivo Excel na resposta
    wb.save(response)

    return response


def visualizar_vaga(request, id):
    if request.method == 'POST':
        if request.user.is_staff:
            gambiarra = {}
            for item in request.POST:
                if item == 'vaga':
                    gambiarra[item] = Cargo.objects.get(nome=request.POST[item]).id
                elif item == 'empresa':
                    gambiarra[item] = Empresa.objects.get(
                        nome=request.POST[item]).id
                else:
                    gambiarra[item] = request.POST[item]
            form = CadastroVagasForm(gambiarra)
            vaga = Vaga_Emprego.objects.get(id=id)
            if form.is_valid():

                form = CadastroVagasForm(gambiarra, instance=vaga)
                form.save()
                return redirect('vagas:vagas')
    else:
        vaga = Vaga_Emprego.objects.get(id=id)
        form = CadastroVagasForm(instance=vaga)

    if request.user.is_staff:
        import datetime
        data_atual = datetime.datetime.now()        
        context = {
            'visualizar': True,
            'mes': data_atual.month,
            'ano': data_atual.year,
            'id': id,
            'tipo_cadastro': '',
            'form': form,
            'hidden': ['user', 'ativo', 'destaque'],
            'cargo': vaga.cargo.nome,
            'empresa': vaga.empresa.nome
        }
    else:
        context = {
            'visualizar': True,
            'id': id,
            'tipo_cadastro': '',
            'form': form,
            'hidden': ['user', 'ativo', 'destaque', 'empresa'],
            'cargo': vaga.cargo.nome,
            'empresa': vaga.empresa.nome
        }

    return render(request, 'vagas/cadastrar_vagaOfertada.html', context)

def vagas(request):    
    vagas = Vaga_Emprego.objects.filter(ativo=True)
    qnt_vagas = len(vagas)
    cont = 0
    for i in vagas:
        cont += i.quantidadeVagas

    context = {
        'vagas': Vaga_Emprego.objects.filter(ativo=True).order_by('cargo__nome'),
        'bairros': Empresa.objects.order_by('bairro').values_list('bairro').distinct(),
        'escolaridades': Escolaridade.objects.all().values(),        
        'qnt_cargos': qnt_vagas,
        'qnt_vagas': cont,        
        'eventos': Slide.objects.all(),
    }
    return render(request, 'vagas/vagas_disponiveis.html', context)

def home(request):
    vagas_destaque = Vaga_Emprego.objects.filter(destaque=True)
    vagas = Vaga_Emprego.objects.filter(ativo=True)
    qnt_vagas = len(vagas)
    cont = 0
    for i in vagas:
        cont += i.quantidadeVagas
    if len(vagas_destaque)>0:
        msg='Vaga em destaque'
    else:
        msg='Vagas em destaques'
    context = {
        'msg': msg,
        'vagas': vagas_destaque,
        'qnt_cargos': qnt_vagas,
        'qnt_vagas': cont,
        'qnt_destaque': len(vagas_destaque),
        'eventos': Slide.objects.all(),
    }
    return render(request, 'vagas/index.html', context)


@login_required
@staff_required
def cadastrar_empresa(request):
    if request.method == 'POST':
        form = Form_Empresa(request.POST)
        if form.is_valid():
            form.save()
            context = {
                'tipo_cadastro': 'Cadastrar',
                'form': Form_Empresa(initial={'user': request.user}),
                'hidden': ['user', 'ativo'],
                'success': [True, 'Empresa cadastrada com sucesso!']
            }
            return render(request, 'vagas/cadastrar_empresa.html', context)
    else:
        form = Form_Empresa(initial={'user': request.user})
    context = {
        'form': form,
        'tipo_cadastro': 'Cadastrar',
    }
    return render(request, 'vagas/cadastrar_empresa.html', context)


@login_required
@staff_required
def alterar_empresa(request, id):
    empresa = Empresa.objects.get(id=id)
    if request.method == 'POST':
        form = Form_Empresa(request.POST, instance=empresa)
        if form.is_valid():
            form.save()
            context = {
                'tipo_cadastro': 'Alterar',
                'form': Form_Empresa(initial={'user': request.user}),
                'hidden': ['user', 'ativo'],
                'success': [True, 'Vaga alterada com sucesso!']
            }
            return redirect('vagas:empresas')
    else:

        form = Form_Empresa(instance=empresa)
    context = {
        'form': form,
        'tipo_cadastro': 'Alterar',
    }
    return render(request, 'vagas/cadastrar_empresa.html', context)


@login_required
@staff_required
def cadastrar_cargo(request):
    if request.method == 'POST':
        form = Form_Cargo(request.POST)
        if form.is_valid():
            form.save()
            context = {
                'tipo_cadastro': 'Cadastrar',
                'form': Form_Cargo(initial={'user': request.user}),
                'hidden': ['user', 'ativo'],
                'success': [True, 'Vaga cadastrada com sucesso!']
            }
            return render(request, 'vagas/cadastrar_cargo.html', context)
    else:
        form = Form_Cargo(initial={'user': request.user})
    context = {
        'form': form,
        'tipo_cadastro': 'Cadastrar',
    }
    return render(request, 'vagas/cadastrar_cargo.html', context)


@login_required
@staff_required
def alterar_cargo(request, id):
    cargo = Cargo.objects.get(id=id)
    if request.method == 'POST':
        form = Form_Cargo(request.POST, instance=cargo)
        if form.is_valid():
            form.save()
            context = {
                'tipo_cadastro': 'Alterar',
                'form': Form_Cargo(initial={'user': request.user}),
                'hidden': ['user', 'ativo'],
                'success': [True, 'Vaga alterada com sucesso!']
            }
            return redirect('vagas:listar_cargos')
    else:

        form = Form_Cargo(instance=cargo)
    context = {
        'form': form,
        'tipo_cadastro': 'Alterar',
    }
    return render(request, 'vagas/cadastrar_escolaridade.html', context)


@login_required
@staff_required
def cadastrar_escolaridade(request):
    if request.method == 'POST':
        form = Form_Escolaridade(request.POST)
        if form.is_valid():
            form.save()
            context = {
                'tipo_cadastro': 'Cadastrar',
                'form': Form_Escolaridade(initial={'user': request.user}),
                'hidden': ['user', 'ativo'],
                'success': [True, 'Vaga cadastrada com sucesso!']
            }
            return render(request, 'vagas/cadastrar_escolaridade.html', context)
    else:
        form = Form_Escolaridade(initial={'user': request.user})
    context = {
        'form': form,
        'tipo_cadastro': 'Cadastrar',
    }
    return render(request, 'vagas/cadastrar_escolaridade.html', context)


@login_required
@staff_required
def alterar_escolaridade(request, id):
    escolaridade = Escolaridade.objects.get(id=id)
    if request.method == 'POST':
        form = Form_Escolaridade(request.POST, instance=escolaridade)
        if form.is_valid():
            form.save()
            context = {
                'tipo_cadastro': 'Alterar',
                'form': Form_Escolaridade(initial={'user': request.user}),
                'hidden': ['user', 'ativo'],
                'success': [True, 'Vaga alterada com sucesso!']
            }
            return redirect('vagas:escolaridades')
    else:

        form = Form_Escolaridade(instance=escolaridade)
    context = {
        'form': form,
        'tipo_cadastro': 'Alterar',
    }
    return render(request, 'vagas/cadastrar_escolaridade.html', context)


@login_required
@staff_required
def cadastrar_vagaOfertada(request):
    if request.method == 'POST':
        gambiarra = {}
        for item in request.POST:
            if item == 'cargo':
                try:
                    gambiarra[item] = Cargo.objects.get(
                        nome=request.POST[item]).id
                except:
                    gambiarra[item] = request.POST[item]
            elif item == 'empresa':
                try:
                    gambiarra[item] = Empresa.objects.get(
                        nome=request.POST[item]).id
                except:
                    gambiarra[item] = request.POST[item]
            else:
                gambiarra[item] = request.POST[item]
        form = CadastroInternoVagasForm(gambiarra)
        if form.is_valid():
            form.save()
            context = {
                'tipo_cadastro': 'cadastrar',
                'form': CadastroInternoVagasForm(initial={'ativo': True, 'user': request.user}),
                'hidden': ['user', 'ativo'],
                'success': [True, 'Vaga cadastrada com sucesso!']
            }
            return render(request, 'vagas/cadastrar_vagaOfertada.html', context)
    else:
        form = CadastroInternoVagasForm(initial={'ativo': True, 'user': request.user})
    context = {
        'tipo_cadastro': 'cadastrar',
        'form': form,
        'hidden': ['user', 'ativo']
    }
    return render(request, 'vagas/cadastrar_vagaOfertada.html', context)


@login_required
@staff_required
def remover_vaga(request, id):
    if request.method == 'POST':
        try:
            vaga = Vaga_Emprego.objects.get(id=request.POST['remover'])
            vaga.ativo = False
            vaga.save()
            return redirect('vagas:vagas')
        except:
            pass
    context = {
        'id': id,
        'vaga': Vaga_Emprego.objects.get(id=id)
    }
    return render(request, 'vagas/remover_vagaOfertada.html', context)


@login_required
@staff_required
def cadastrar_vaga_emLote(request):
    if request.method == 'POST':
        try:
            empresa = Empresa.objects.get(nome=request.POST['empresa'])
            success = True
        except:
            success = False
        if success:
            form = CadastroInternoVagasForm(
                initial={'ativo': True, 'user': request.user})
            context = {
                'empresa': request.POST['empresa'],
                'tipo_cadastro': 'cadastrar',
                'form': form,
                'hidden': ['user', 'ativo']
            }
            return render(request, 'vagas/cadastrar_vagas_emLote_2.html', context)
    form = CadastroInternoVagasForm(initial={'ativo': True, 'user': request.user})
    context = {
        'tipo_cadastro': 'cadastrar',
        'form': form,
        'hidden': ['user', 'ativo']
    }
    return render(request, 'vagas/cadastrar_vagas_emLote.html', context)


def get_empresa(request):
    try:
        # empresas=Empresa.objects.filter(nome__startswith=request.GET.get('nome')).order_by('nome')
        empresas = Empresa.objects.filter(
            nome__icontains=request.GET.get('empresa')).order_by('nome')
    except Exception as E:
        empresas = None
    context = {
        'results': empresas,
    }
    return render(request, 'vagas/resultEmpresaSearchs.html', context)


def get_cargo(request):
    try:
        # empresas=Empresa.objects.filter(nome__startswith=request.GET.get('nome')).order_by('nome')
        cargos = Cargo.objects.filter(
            nome__icontains=request.GET.get('vaga')).order_by('nome')
    except Exception as E:
        cargos = None
    context = {
        'results': cargos,
    }
    return render(request, 'vagas/resultVagaSearchs.html', context)




@login_required
@staff_required
def alterar_vaga(request, id):
    if request.method == 'POST':
        gambiarra = {}
        for item in request.POST:
            if item == 'cargo':
                try:
                    gambiarra[item] = Cargo.objects.get(
                        nome=request.POST[item]).id
                except:
                    gambiarra[item] = request.POST[item]
            elif item == 'empresa':
                try:
                    gambiarra[item] = Empresa.objects.get(
                        nome=request.POST[item]).id
                except:
                    gambiarra[item] = request.POST[item]
            else:
                gambiarra[item] = request.POST[item]
        form = CadastroInternoVagasForm(gambiarra)
        vaga = Vaga_Emprego.objects.get(id=id)
        if form.is_valid():

            form = CadastroVagasForm(gambiarra, instance=vaga)
            form.save()
            return redirect('vagas:vagas')
    else:
        vaga = Vaga_Emprego.objects.get(id=id)
        form = CadastroInternoVagasForm(instance=vaga)

    context = {
        'id': id,
        'tipo_cadastro': 'Alterar',
        'form': form,
        'hidden': ['user', 'ativo'],
        'cargo': vaga.cargo.nome,
        'empresa': vaga.empresa.nome
    }
    return render(request, 'vagas/cadastrar_vagaOfertada.html', context)




@staff_required
def empresas(request):
    context = {
        'empresas': Empresa.objects.all()
    }
    return render(request, 'vagas/listar_empresas.html', context)

@staff_required
def escolaridades(request):
    context = {
        'escolaridades': Escolaridade.objects.all()
    }
    return render(request, 'vagas/listar_escolaridade.html', context)

@staff_required
def listar_cargos(request):
    context = {
        'vagas': Cargo.objects.all()
    }
    return render(request, 'vagas/listar_cargos.html', context)

@staff_required
def imprimir_vagas(request):
    vagas = Vaga_Emprego.objects.filter(ativo=True).order_by('cargo__nome')
    cont = 0
    for i in vagas:
        cont += i.quantidadeVagas

    context = {
        'vagas': vagas,
        'total': cont
    }
    return render(request, 'vagas/imprimir_vagas.html', context)


@xframe_options_exempt
def vagas_table(request):
    context = {
        'vagas': Vaga_Emprego.objects.filter(ativo=True)
    }
    return render(request, 'vagas/vagas.html', context)


def login_view(request):
    if request.user.is_authenticated:
        return redirect('/')
    if request.method == 'POST':
        # Abaixo recebemos a validação da API do Google do reCAPTCHA
        ''' Begin reCAPTCHA validation '''
        recaptcha_response = request.POST.get('g-recaptcha-response')
        data = {
            'secret': '6LdiIsweAAAAADv7tYKHZ1fCP4pi6FwIZTw4X4Rl',
            'response': recaptcha_response
        }
        r = requests.post(
            'https://www.google.com/recaptcha/api/siteverify', data=data)
        result = r.json()
        ''' End reCAPTCHA validation '''

        # Se o reCAPTCHA garantir que o usuário é um robô
        if result['success']:
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                try:
                    return redirect(request.GET['next'])
                except:
                    return redirect('vagas:home')
            else:
                context = {
                    'error': True,
                }
                return render(request, 'registration/login.html', context)
        else:
            context = {
                'error2': True,
            }
            return render(request, 'registration/login.html', context)
    return render(request, 'registration/login.html')


def encaminhar(request, id):
    today = date.today()
    vaga = Vaga_Emprego.objects.get(id=id)
    if request.method == 'POST':
        context = {
            'vaga': vaga,
            'date': today,
            'candidato': {'nome': request.POST['nome']}
        }
        return render(request, 'vagas/encaminhar.html', context)

    return redirect('vagas:encaminhamento', id)


def encaminhamento(request, id, user_id=0):
    candidato = Candidato.objects.get(id=id)
    if str(int(user_id)) != str(int(0)):
        if request.user.is_staff:
            user = User.objects.get(id=user_id)
        else:
            user = False    
    else:
        user = False

    from datetime import date
    today = date.today()
    context = {
        'vaga': candidato.vaga,
        'date': today,
        'candidato': candidato,
        'sistema': True,
        'user': user
    }
    if request.user.is_staff:
        return render(request, 'vagas/encaminhar.html', context)
    return render(request, 'vagas/encaminhamento_online.html', context)


def gera_encaminhamento_to_pdf(request, id, user_id=0):
    try:
        url_pdf = '/home/casa_do_trabalhador/site/balcao_de_emprego/vagas/static/pdf/' + \
            str(id)+'.pdf'
        # url_pdf='/home/eduardo/projects/casadotrabalhador/vagas/static/pdf/'+id+'.pdf'
        pdfkit.from_url('https://casadotrabalhador.pmnf.rj.gov.br/visualizar-vaga/alt0x' +
                        str(id)+'0'+str(user_id)+'01/encaminhamento', url_pdf)
        # pdfkit.from_url('http://localhost:8000/visualizar-vaga/alt0x'+str(id)+'0'+str(user_id)+'01/encaminhamento', url_pdf)

        context = {
            'pdf': url_pdf
        }
        try:
            return FileResponse(open(url_pdf, 'rb'), content_type='application/pdf')
        except Exception as E:
            raise Http404()
    except Exception as E:
        return redirect('/')


def candidatarse(request, id):
    if request.user.is_staff:
        form = Form_Candidato(initial={'vaga': id, 'candidato_online': False})
    else:
        form = Form_Candidato(initial={'vaga': id, 'candidato_online': True})

    if request.method == 'POST':
        form = Form_Candidato(request.POST)
        if form.is_valid():

            try:
                cpf = validate_CPF(request.POST['cpf'])
                candidato = Candidato.objects.get(cpf=cpf, vaga_id=id)
                form = Form_Candidato(request.POST, instance=candidato)

            except Exception as e:
                pass

            candidato = form.save()
            # return render(request, 'vagas/encaminhar.html', context)
            if request.user.is_staff:
                candidato.funcionario_encaminhamento = request.user
                candidato.dt_atualizacao = datetime.today()
                candidato.save()

                return redirect('vagas:encaminhamento', id=candidato.id, user_id=request.user.id)
            return redirect('vagas:encaminhamento', id=candidato.id, user_id=0)

    context = {
        'id': id,
        'form': form
    }
    return render(request, 'vagas/candidatarse.html', context)


@login_required
@staff_required
def candidatosporvaga(request, id, mes, ano):
    candidatos = Candidato.objects.filter(vaga=id).order_by('dt_inclusao')

    if mes and ano:
        date = datetime(int(ano), int(mes), 1)

        if date.month == 12:
            _, last_day = calendar.monthrange(date.year + 1, 1)
            end_date = datetime(date.year + 1, 1, 1) + timedelta(days=last_day - 1)
        else:
            _, last_day = calendar.monthrange(date.year, date.month + 1)
            end_date = date + timedelta(days=last_day)

        candidatos = candidatos.filter(dt_inclusao__range=(date, end_date))


    paginator = Paginator(candidatos, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    page_obj.page_range = paginator.page_range

    context = {
        'candidatos': page_obj,
        'id': id,
        'mes': mes,
        'ano': ano
    }

    return render(request, 'vagas/listar_candidatos.html', context)

def infoempresa(request):
    empresas = Empresa.objects.all()
    context={
        'empresas': empresas
    }
    return render(request, 'vagas/infoempresa.html', context)

def infoempresa_download(request, id):
    empresa = get_object_or_404(Empresa, id=id)

    # Obtém todos os candidatos encaminhados para a empresa específica
    candidatos = Candidato.objects.filter(vaga__empresa=empresa)

    # Cria um novo workbook e uma planilha
    wb = Workbook()
    ws = wb.active

    # Configuração do cabeçalho do arquivo Excel
    ws.append(["Informações da Empresa: " + empresa.nome + " - " + empresa.cnpj +" Data emissão: " + str(date.today())])
    header = ["Nome", "CPF", "Data de Nascimento", "Sexo", "E-mail", "Celular", "Bairro", "Escolaridade", "Online", "Data de Inclusão"]

    # Adiciona o cabeçalho à planilha
    ws.append(header)

    # Adiciona os dados dos candidatos à planilha
     # Mantém o controle dos CPFs já adicionados
    cpf_set = set()

    # Adiciona os dados dos candidatos à planilha, evitando CPFs duplicados
    for candidato in candidatos:
        if candidato.cpf not in cpf_set:
            data_row = [candidato.nome, candidato.cpf, candidato.data_nascimento, candidato.get_sexo_display(), candidato.email, candidato.celular, candidato.bairro, candidato.escolaridade.nome, candidato.candidato_online, candidato.dt_inclusao]
            ws.append(data_row)
            cpf_set.add(candidato.cpf)

    # Ajusta o alinhamento do texto na planilha
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')


    # Cria a resposta HTTP com o conteúdo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Candidatos_{empresa.nome}.xlsx'

    # Salva o workbook na resposta HTTP
    wb.save(response)

    return response

@login_required
@staff_required
def pesquisar_candidatos(request):
    context = {}
    return render(request, 'vagas/pesquisar_candidatos.html', context)


@login_required
def get_candidatos(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode("utf-8"))
        cpf = ''
        nome = ''
        try:
            cpf = data['cpf']
        except:
            nome = data['nome']
        
        if cpf:
            try:
                # empresas=Empresa.objects.filter(nome__startswith=request.GET.get('nome')).order_by('nome')
                candidatos = Candidato.objects.filter(cpf=cpf)
            except Exception as E:
                candidatos = None
        elif nome:
            try:
                # empresas=Empresa.objects.filter(nome__startswith=request.GET.get('nome')).order_by('nome')
                candidatos = Candidato.objects.filter(
                    nome__icontains=nome).order_by('nome')
            except Exception as E:
                candidatos = None

        else:
            print('dados informados não são os esperados')
            candidatos = None

        context = {
            'candidatos': candidatos,
        }

    return render(request, 'vagas/pesquisar_candidatos_result.html', context)


@login_required
@staff_required
def visualizar_candidato(request, id):
    candidato = Candidato.objects.get(id=id)
    context = {
        'candidato': candidato,
        'form': Form_Candidato(instance=candidato)
    }
    return render(request, 'vagas/pesquisar_candidatos_visualizar.html', context)


@login_required
@staff_required
def vagascomcandidatos(request):    
    buscar = False
    context = {}

    if request.method == 'POST':

        month=request.POST['mes']
        year=request.POST['ano']

        buscar = True
        date = datetime(int(year), int(month), 1)

        if date.month == 12:
            _, last_day = calendar.monthrange(date.year + 1, 1)
            end_date = datetime(date.year + 1, 1, 1) + timedelta(days=last_day - 1)
        else:
            _, last_day = calendar.monthrange(date.year, date.month + 1)
            end_date = date + timedelta(days=last_day)

        candidatos_interval = Candidato.objects.filter(dt_inclusao__range=(date, end_date))

        vagas_com_candidatos = []

        vagas = Vaga_Emprego.objects.all()
        for vaga in vagas:
            candidatos = candidatos_interval.filter(vaga=vaga.id).count()
            if candidatos > 0:
                vagas_com_candidatos.append({'informacao': vaga, 'total': candidatos})

        paginator = Paginator(vagas_com_candidatos, 100)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        

        queryOnline=f'''
        SELECT DISTINCT id, cpf
        FROM vagas_candidato 
        WHERE MONTH(dt_inclusao)='{month}' 
        AND YEAR(dt_inclusao)='{year}' AND candidato_online=1;'''

        queryBalcao=f'''
        SELECT DISTINCT id, cpf
        FROM vagas_candidato 
        WHERE MONTH(dt_inclusao)='{month}' 
        AND YEAR(dt_inclusao)='{year}' AND candidato_online=0;'''             
        
        context = {
            'vagas': page_obj,
            'balcao':len(list(Vaga_Emprego.objects.raw(queryBalcao))),
            'online': len(list(Vaga_Emprego.objects.raw(queryOnline))),
            'buscar': buscar,
            'ano': year,
            'mes': month
        }
   
    return render(request, 'vagas/vagas_com_candidatos.html', context)


@login_required
@staff_required
def candidatosporfuncionario(request):
    usuarios = User.objects.filter(groups__name='atendente')
    lista = []
    month = ''
    year = ''
    
    if request.method == 'POST':

        month=request.POST['mes']
        year=request.POST['ano']

        date = datetime(int(year), int(month), 1)

        if date.month == 12:
            _, last_day = calendar.monthrange(date.year + 1, 1)
            end_date = datetime(date.year + 1, 1, 1) + timedelta(days=last_day - 1)
        else:
            _, last_day = calendar.monthrange(date.year, date.month + 1)
            end_date = date + timedelta(days=last_day)

        candidatos_interval = Candidato.objects.filter(dt_inclusao__range=(date, end_date))

    else: 
        candidatos_interval = Candidato.objects.all()

    for i in usuarios:
        lista.append([i.first_name, len(
            candidatos_interval.filter(funcionario_encaminhamento=i)), i.id])
        
    context = {
        'lista': lista,
        'mes': month,
        'ano': year
    }

    return render(request, 'vagas/candidatos_por_funcionarios.html', context)


@login_required
@staff_required
def funcionario_encaminhados(request, id):
    candidatos = Candidato.objects.filter(funcionario_encaminhamento=id)
    paginator = Paginator(candidatos, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    page_obj.page_range = paginator.page_range

    context = {
        'candidatos': page_obj,
        'fulano': User.objects.get(id=id).first_name,
        'id': id
    }

    return render(request, 'vagas/candidatos_por_funcionarios_detalhe.html', context)


@login_required
def sair(request):
    if request.user.is_authenticated:
        logout(request)
        return redirect('vagas:home')
    else:
        return redirect('/accounts/login')


@login_required
@staff_required
def painel_administrativo(request):
    return render(request, 'vagas/painel_administrativo.html')


@login_required
@staff_required
def painel_administrativo_excluir_cpf(request):
    return render(request, 'vagas/painel_administrativo_excluir_cpf.html')


@login_required
@staff_required
def excluir_cpf(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode("utf-8"))
        if data['step'] == 0:
            cpf = validate_CPF(data['cpf'])
            potencialmente_excluidos = Candidato.objects.filter(
                cpf=cpf).count()

            return JsonResponse({'qnt_excluidos': potencialmente_excluidos, 'cpf': cpf, 'step': 1})

        elif data['step'] == 1:
            cpf = validate_CPF(data['cpf'])
            excluidos = Candidato.objects.filter(cpf=cpf).delete()
            return JsonResponse({'qnt_excluidos': excluidos[0], 'cpf': cpf, 'step': 1})


@login_required
@staff_required
def indicadores(request):

    timezone.activate(settings.TIME_ZONE)
    start_date = date(2022, 9, 1)
    end_date = date.today()

    top_x = 11

    # -------------------- #

    vagas_emprego_faixa = Vaga_Emprego.objects.filter(dt_inclusao__gte = start_date, dt_inclusao__lt = end_date)
    cargos = Cargo.objects.all()
    escolaridades = Escolaridade.objects.all()
    empresas = Empresa.objects.all()

    # -------------------- #

    cargos_ofertados = []
    escolaridades_quantidades = []
    vagas_por_empresa = []
    candidatos_por_mes = []
    vagas_cadastradas_por_mes = []
    relacao_online_balcao = []

    # -------------------- #

    for cargo in cargos:
        total = vagas_emprego_faixa.filter(cargo=cargo).aggregate(Sum('quantidadeVagas'))['quantidadeVagas__sum']
        if total == None:
            total = 0

        cargos_ofertados.append({'nome': cargo.nome, 'total': total})
    
    cargos_ofertados = sorted(cargos_ofertados, key=lambda x: x['total'], reverse=True)

    # ------------------------- #

    for escolaridade in escolaridades:
        total = Candidato.objects.filter(dt_inclusao__gte = start_date, dt_inclusao__lt = end_date, escolaridade=escolaridade).values('email').distinct().count()
        if total == None:
            total = 0

        escolaridades_quantidades.append({'nome': escolaridade.nome, 'total': total})

    # --------------------------- #
    
    delta = relativedelta(months=1)

    while start_date <= end_date:
        next_month = start_date + delta
        total = Candidato.objects.filter(dt_inclusao__gte = start_date, dt_inclusao__lt = next_month).values('email').distinct().count()
        if total == None:
            total = 0

        candidatos_por_mes.append({'nome': start_date, 'total': total})
        start_date += delta

    # ----------------------------- #

    start_date = date(2022, 9, 1)

    while start_date <= end_date:
        next_month = start_date + delta
        total = vagas_emprego_faixa.filter(dt_inclusao__gte = start_date, dt_inclusao__lt = next_month).aggregate(Sum('quantidadeVagas'))['quantidadeVagas__sum']
        if total == None:
            total = 0

        vagas_cadastradas_por_mes.append({'nome': start_date, 'total': total})
        start_date += delta

    # ----------------------------- #

    for empresa in empresas:
        total = vagas_emprego_faixa.filter(empresa=empresa).aggregate(Sum('quantidadeVagas'))['quantidadeVagas__sum']

        if total == None:
            total = 0

        vagas_por_empresa.append({'nome': empresa.nome, 'total': total})

    vagas_por_empresa = sorted(vagas_por_empresa, key=lambda x: x['total'], reverse=True)

    # ----------------------------- #

    start_date = date(2022, 9, 1)

    while start_date <= end_date:
        next_month = start_date + delta
        candidatos = Candidato.objects.filter(dt_inclusao__gte = start_date, dt_inclusao__lt = next_month).values('email').distinct()
        
        candidatos_online = candidatos.filter(candidato_online = True).count()
        candidatos_balcao = candidatos.filter(candidato_online = False).count()

        if total == None:
            total = 0

        relacao_online_balcao.append({'nome': start_date, 'online': candidatos_online, 'balcao': candidatos_balcao})
        start_date += delta

        
    context = {
        'top_x': top_x,
        'top_cargos_ofertados': cargos_ofertados[:top_x],
        'escolaridades': escolaridades_quantidades,
        'candidatos_por_mes': candidatos_por_mes,
        'vagas_por_empresa': vagas_por_empresa[:top_x],
        'vagas_cadastradas_por_mes': vagas_cadastradas_por_mes,
        'relacao_online_balcao': relacao_online_balcao
    }

    return render(request, 'vagas/indicadores.html', context)

@login_required
@staff_required
def emails(request):

    context = {
        'buscar': False
    }

    if request.method == 'POST':

        month=request.POST['mes']
        year=request.POST['ano']

        date = datetime(int(year), int(month), 1)

        if date.month == 12:
            _, last_day = calendar.monthrange(date.year + 1, 1)
            end_date = datetime(date.year + 1, 1, 1) + timedelta(days=last_day - 1)
        else:
            _, last_day = calendar.monthrange(date.year, date.month + 1)
            end_date = date + timedelta(days=last_day)

        candidatos = Candidato.objects.filter(dt_inclusao__range=(date, end_date), email__isnull = False).exclude(email__exact='').values('email', 'nome').distinct()

        context = {
            'data': date,
            'mes': month,
            'ano': year,
            'candidatos': candidatos,
            'buscar': True,
        }

    return render(request, 'vagas/emails.html', context) 

@login_required
@staff_required
def download_emails(request, month, year):

    date = datetime(int(year), int(month), 1)

    if date.month == 12:
        _, last_day = calendar.monthrange(date.year + 1, 1)
        end_date = datetime(date.year + 1, 1, 1) + timedelta(days=last_day - 1)
    else:
        _, last_day = calendar.monthrange(date.year, date.month + 1)
        end_date = date + timedelta(days=last_day)

    candidatos = Candidato.objects.filter(dt_inclusao__range=(date, end_date), email__isnull = False).exclude(email__exact='').values('nome','email').distinct()

    context = {
        'listas': candidatos,
    }

    return render(request, 'vagas/email_csv.html', context) 

# Função temporária
def manutencao(request):
    return render(request, 'vagas/manutencao.html')

def meus_encaminhamentos(request):    
    try:
        pessoa = Pessoa.objects.get(user=request.user)
        candidato_encaminhamentos = Candidato.objects.filter(cpf=pessoa.cpf, vaga__ativo=True)
    except:
        candidato_encaminhamentos = None
        
    context = {
       'encaminhamentos': candidato_encaminhamentos
    }
    return render(request, 'vagas/meus_encaminhamentos.html', context)
